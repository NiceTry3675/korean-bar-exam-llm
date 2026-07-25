"""
Excel-JSON 양방향 동기화 스크립트

기능:
- Export: Excel 데이터를 results_verified.json 형식으로 내보내기
- Import: results_verified.json을 읽어 Excel에 새 모델 컬럼 추가

사용법:
    python sync_data.py export --sheet 공법 --model "Claude Fable 5 (max)"
    python sync_data.py export --sheet 공법 --all-models
    python sync_data.py export --all-sheets                      # 모든 과목을 하나의 JSON 배열로 출력
    python sync_data.py export --all-sheets --output bar_exam_all_results.json
    python sync_data.py import --json problems/bar-exam-15/public-law/results_verified.json
    python sync_data.py import --all
    python sync_data.py list
    python sync_data.py validate
"""

import json
import argparse
import math
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from copy import deepcopy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import Color


REFUSAL_ANSWER = -2
NO_ANSWER = -1
REFUSAL_MARKERS = {"-2", "(검열)", "검열", "Refusal", "refusal"}
PARSE_FAILED_MARKERS = {"(파싱 실패)", "파싱 실패", "parse_failed"}
DEFAULT_EXCEL_PATH = Path('제15회 변호사시험 LLM 풀이.xlsx')
DEFAULT_HARD_EXCEL_PATH = Path('제15회 변호사시험 LLM 풀이 hard.xlsx')
DEFAULT_BENCHMARK_ID = 'bar-exam-15'
DEFAULT_REGISTRY_PATH = Path('benchmarks/registry.json')
# 모델 결과가 아니라 채점 구조를 담는 컬럼 머리글
STRUCTURE_COLUMN_HEADERS = {'문항 번호', '정답', '배점', 'nan', ''}


def _normalize_numeric_value(value: Any) -> Optional[float]:
    """
    @brief Excel/JSON 숫자를 소수점 손실 없이 정규화한다.

    @param value 변환할 값
    @return 정수 또는 실수 값, 변환할 수 없으면 None
    """
    try:
        number = float(value)
    except (ValueError, TypeError):
        return None
    return int(number) if number.is_integer() else number


def normalize_answer_value(answer):
    """답안 값을 JSON 채점용 숫자와 상태로 변환한다."""
    answer_text = str(answer).strip() if answer is not None else ''
    if answer_text in PARSE_FAILED_MARKERS:
        return None, 'parse_failed'
    if answer is None or answer == '' or str(answer).strip() == '':
        return NO_ANSWER, 'no_answer'

    if answer == NO_ANSWER or answer_text in {'-1', '(포기)'}:
        return NO_ANSWER, 'no_answer'
    if answer_text in REFUSAL_MARKERS:
        return REFUSAL_ANSWER, 'refusal'

    try:
        return int(answer), 'answered'
    except (ValueError, TypeError):
        return None, 'parse_failed'


def _create_hard_excel_template(source_path: Path, target_path: Path):
    """
    @brief 기존 Excel에서 모델 결과 컬럼만 제거한 hard 전용 템플릿을 만든다.

    문항 번호·정답·배점 등 채점 구조 컬럼은 위치와 무관하게 보존한다.

    @param source_path 원본 워크북 경로
    @param target_path 생성할 템플릿 경로
    @throws FileNotFoundError 원본 워크북이 없을 때
    """
    if not source_path.exists():
        raise FileNotFoundError(f"hard Excel 템플릿 원본을 찾을 수 없습니다: {source_path}")

    workbook = load_workbook(source_path)
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        header_row = None
        for row_idx in range(1, min(6, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and '문항 번호' in str(cell_value):
                header_row = row_idx
                break

        if not header_row:
            continue

        model_columns = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if not cell_value:
                continue
            header = str(cell_value).strip()
            if header in STRUCTURE_COLUMN_HEADERS or 'Unnamed' in header:
                continue
            model_columns.append(col_idx)

        for col_idx in sorted(model_columns, reverse=True):
            ws.delete_cols(col_idx)

    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_path)
    print(f"hard Excel 템플릿 생성 완료: {target_path}")


class PathMapper:
    """Excel 시트명과 JSON 경로 간 벤치마크별 매핑"""

    def __init__(self, base_dir: Path = None,
                 benchmark_id: str = DEFAULT_BENCHMARK_ID,
                 registry_path: Path = DEFAULT_REGISTRY_PATH):
        self.base_dir = base_dir or Path('.')
        self.benchmark_id = benchmark_id
        self.registry_path = Path(registry_path)
        self.benchmark = self._load_benchmark()
        self.SHEET_TO_JSON = {
            section['sheet']: section['problemDir']
            for section in self.benchmark.get('sections', [])
            if section.get('sheet') and section.get('problemDir')
        }
        self._section_configs = {
            section['sheet']: section
            for section in self.benchmark.get('sections', [])
            if section.get('sheet')
        }

        # 역방향 매핑 생성
        self._json_to_sheet = {v: k for k, v in self.SHEET_TO_JSON.items()}

    def _load_benchmark(self) -> Dict:
        """
        @brief 추적되는 레지스트리에서 현재 벤치마크 설정을 읽는다.

        @return 선택한 벤치마크 설정
        @throws FileNotFoundError 레지스트리 파일이 없을 때
        @throws ValueError 레지스트리는 존재하지만 벤치 ID가 없을 때
        """
        registry_file = self.registry_path
        if not registry_file.is_absolute():
            registry_file = self.base_dir / registry_file
        if not registry_file.exists():
            raise FileNotFoundError(f"벤치마크 레지스트리를 찾을 수 없습니다: {registry_file}")

        with open(registry_file, 'r', encoding='utf-8') as file:
            registry = json.load(file)

        benchmarks = registry.get('benchmarks', [])
        benchmark = next(
            (item for item in benchmarks if item.get('id') == self.benchmark_id),
            None,
        )
        if benchmark is None:
            available = ', '.join(item.get('id', '?') for item in benchmarks)
            raise ValueError(
                f"알 수 없는 벤치마크: {self.benchmark_id} "
                f"(사용 가능: {available or '없음'})"
            )
        return benchmark

    def sheet_to_json_path(self, sheet_name: str) -> Optional[Path]:
        """시트 이름으로 JSON 폴더 경로 반환"""
        if sheet_name in self.SHEET_TO_JSON:
            return self.base_dir / self.SHEET_TO_JSON[sheet_name]
        return None

    def json_to_sheet_name(self, json_path: Path) -> Optional[str]:
        """JSON 경로로 시트 이름 반환"""
        # 경로 정규화
        rel_path = str(json_path).replace('\\', '/')
        # results_verified.json 제거
        if rel_path.endswith('/results_verified.json'):
            rel_path = rel_path[:-len('/results_verified.json')]
        elif rel_path.endswith('results_verified.json'):
            rel_path = str(Path(rel_path).parent).replace('\\', '/')

        # 매핑 검색
        for json_rel, sheet in self._json_to_sheet.items():
            if rel_path.endswith(json_rel) or json_rel in rel_path:
                return sheet
        return None

    def get_all_sheets(self) -> List[str]:
        """모든 시트 이름 반환"""
        return list(self.SHEET_TO_JSON.keys())

    def get_section_config(self, sheet_name: str) -> Dict:
        """
        @brief 시트에 대응하는 레지스트리 섹션 설정을 반환한다.

        @param sheet_name Excel 시트 이름
        @return 섹션 설정, 등록되지 않은 시트면 빈 딕셔너리
        """
        return self._section_configs.get(sheet_name, {})

    def get_mode_config(self, hard_mode: bool = False) -> Dict:
        """
        @brief 현재 벤치마크의 실행 모드 설정을 반환한다.

        @param hard_mode 과목 일괄(hard) 모드 여부
        @return 모드 설정
        """
        mode_name = 'hard' if hard_mode else 'default'
        return self.benchmark.get('modes', {}).get(mode_name, {})

    def load_questions(self, sheet_name: str) -> Dict:
        """
        @brief 공개 정답을 정본으로 읽고, 없는 레거시 섹션만 로컬 데이터를 사용한다.

        @param sheet_name Excel 시트 이름
        @return 기존 questions.json 호환 객체
        @throws FileNotFoundError 로컬·공개 메타데이터가 모두 없을 때
        """
        section_config = self.get_section_config(sheet_name)
        metadata_path = section_config.get('metadataPath')
        if metadata_path:
            metadata_file = Path(metadata_path)
            if not metadata_file.is_absolute():
                metadata_file = self.base_dir / metadata_file
            if not metadata_file.exists():
                raise FileNotFoundError(f"공개 정답 메타데이터를 찾을 수 없습니다: {metadata_file}")
            with open(metadata_file, 'r', encoding='utf-8') as file:
                metadata = json.load(file)
            subject_data = next(
                (
                    item for item in metadata.get('subjects', [])
                    if item.get('sheet_name') == sheet_name
                    or item.get('id') == section_config.get('id')
                    or item.get('subject') == section_config.get('subject')
                ),
                None,
            )
            if not subject_data:
                raise ValueError(
                    f"공개 정답 메타데이터에 시트가 없습니다: {metadata_file} / {sheet_name}"
                )
            return {
                'benchmark_id': metadata.get('benchmark_id'),
                'subject': subject_data.get('subject', sheet_name),
                'section': subject_data.get(
                    'section', subject_data.get('subject', sheet_name)
                ),
                'questions': subject_data.get('questions', []),
            }

        json_path = self.sheet_to_json_path(sheet_name)
        local_file = json_path / 'questions.json' if json_path else None
        if local_file and local_file.exists():
            with open(local_file, 'r', encoding='utf-8') as file:
                return json.load(file)

        expected_paths = [str(path) for path in (local_file, metadata_path) if path]
        raise FileNotFoundError(
            f"questions.json을 찾을 수 없습니다: {', '.join(expected_paths) or sheet_name}"
        )

    def get_subject_section(self, sheet_name: str) -> Tuple[str, str]:
        """시트 이름에서 과목, 섹션 추출"""
        section_config = self.get_section_config(sheet_name)
        if section_config:
            return (
                section_config.get('subject', sheet_name),
                section_config.get('section', section_config.get('subject', sheet_name)),
            )
        if '-' in sheet_name:
            parts = sheet_name.split('-', 1)
            return parts[0].strip(), parts[1].strip()
        return sheet_name.strip(), sheet_name.strip()


class ModelNameMapper:
    """모델 이름 매핑 (JSON 이름 <-> Excel 이름)"""

    def __init__(self, mapping_file: Path = None):
        self.mapping_file = mapping_file or Path('model_mapping.json')
        self.mapping = {}  # JSON name -> Excel name
        self._load_mapping()

    def _load_mapping(self):
        """매핑 파일 로드"""
        if self.mapping_file.exists():
            with open(self.mapping_file, 'r', encoding='utf-8') as f:
                self.mapping = json.load(f)

    def save_mapping(self):
        """매핑 파일 저장"""
        with open(self.mapping_file, 'w', encoding='utf-8') as f:
            json.dump(self.mapping, f, ensure_ascii=False, indent=2)

    def json_to_excel(self, json_name: str) -> str:
        """JSON 모델 이름을 Excel 컬럼 이름으로 변환"""
        return self.mapping.get(json_name, json_name)

    def excel_to_json(self, excel_name: str) -> str:
        """Excel 컬럼 이름을 JSON 모델 이름으로 변환"""
        # 역방향 검색
        for json_name, mapped_excel in self.mapping.items():
            if mapped_excel == excel_name:
                return json_name
        return excel_name

    def add_mapping(self, json_name: str, excel_name: str):
        """새 매핑 추가"""
        self.mapping[json_name] = excel_name


class ExcelHandler:
    """Excel 파일 읽기/쓰기"""

    def __init__(self, excel_path: Path, path_mapper: PathMapper = None):
        self.excel_path = Path(excel_path)
        self.path_mapper = path_mapper or PathMapper(base_dir=Path('.'))
        self.workbook = None
        self._header_row_cache = {}

    def _load_workbook(self):
        """워크북 로드 (lazy loading)"""
        if self.workbook is None:
            self.workbook = load_workbook(self.excel_path)

    def get_sheet_names(self) -> List[str]:
        """모든 시트 이름 반환"""
        self._load_workbook()
        return self.workbook.sheetnames

    def _find_header_row(self, sheet_name: str) -> int:
        """헤더 행 번호 찾기 (1-based)"""
        if sheet_name in self._header_row_cache:
            return self._header_row_cache[sheet_name]

        self._load_workbook()
        ws = self.workbook[sheet_name]

        for row_idx in range(1, min(6, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and '문항 번호' in str(cell_value):
                self._header_row_cache[sheet_name] = row_idx
                return row_idx

        raise ValueError(f"'{sheet_name}' 시트에서 헤더 행을 찾을 수 없습니다.")

    def _find_score_row(self, sheet_name: str) -> int:
        """총점 행 번호 찾기 (1-based)"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)

        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and str(cell_value).strip() in ['총점', '총합', '점수']:
                return row_idx

        raise ValueError(f"'{sheet_name}' 시트에서 총점 행을 찾을 수 없습니다.")

    def get_model_columns(self, sheet_name: str) -> Dict[str, int]:
        """모델 컬럼 이름과 열 번호 반환 (1-based)"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)

        models = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value:
                col_str = str(cell_value).strip()
                # 불필요한 컬럼 제외
                if col_str in STRUCTURE_COLUMN_HEADERS:
                    continue
                if 'Unnamed' in col_str:
                    continue
                models[col_str] = col_idx

        return models

    def get_model_answers(self, sheet_name: str, model_name: str) -> Dict[int, any]:
        """특정 모델의 문항별 답 추출"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)
        score_row = self._find_score_row(sheet_name)

        model_columns = self.get_model_columns(sheet_name)
        if model_name not in model_columns:
            raise ValueError(f"'{model_name}' 모델을 '{sheet_name}' 시트에서 찾을 수 없습니다.")

        col_idx = model_columns[model_name]
        answers = {}

        for row_idx in range(header_row + 1, score_row):
            q_num = ws.cell(row=row_idx, column=1).value
            answer = ws.cell(row=row_idx, column=col_idx).value

            if q_num is not None:
                try:
                    q_num = int(q_num)
                    answers[q_num] = answer
                except (ValueError, TypeError):
                    pass

        return answers

    def calculate_score_from_answers(self, sheet_name: str,
                                     answers: Dict[int, Any]) -> float:
        """문항별 답안으로 총점 계산"""
        self._load_workbook()
        correct_answers = self._get_correct_answers(sheet_name)
        questions_data = self._load_questions_for_sheet(sheet_name)
        points_by_question = {q['number']: q['points'] for q in questions_data['questions']}

        score = 0
        for q_num, correct in correct_answers.items():
            answer = answers.get(q_num)
            normalized_answer, _ = normalize_answer_value(answer)
            try:
                if normalized_answer == int(correct):
                    score += points_by_question.get(q_num, 0)
            except (ValueError, TypeError):
                continue

        return score

    def _load_questions_for_sheet(self, sheet_name: str) -> Dict:
        """시트에 대응하는 questions.json 로드"""
        return self.path_mapper.load_questions(sheet_name)

    def get_model_score(self, sheet_name: str,
                        model_name: str) -> Optional[float]:
        """특정 모델의 총점 반환"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        score_row = self._find_score_row(sheet_name)

        model_columns = self.get_model_columns(sheet_name)
        if model_name not in model_columns:
            return None

        col_idx = model_columns[model_name]
        score = ws.cell(row=score_row, column=col_idx).value

        return _normalize_numeric_value(score)

    def get_max_score(self, sheet_name: str) -> float:
        """만점 반환 (정답 열의 총점)"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)
        score_row = self._find_score_row(sheet_name)

        # 정답 열 찾기
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value and str(cell_value).strip() == '정답':
                max_score = ws.cell(row=score_row, column=col_idx).value
                normalized = _normalize_numeric_value(max_score)
                if normalized is not None:
                    return normalized

        configured_max = self.path_mapper.get_section_config(sheet_name).get('maxScore')
        normalized = _normalize_numeric_value(configured_max)
        if normalized is not None:
            return normalized

        return 100  # 기본값

    def _find_answer_column(self, sheet_name: str) -> int:
        """정답 열 번호 찾기 (1-based)"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)

        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value and str(cell_value).strip() == '정답':
                return col_idx

        return 2  # 기본값

    def _get_correct_answers(self, sheet_name: str) -> Dict[int, any]:
        """정답 데이터 추출 {문항번호: 정답}"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)
        score_row = self._find_score_row(sheet_name)
        answer_col = self._find_answer_column(sheet_name)

        correct_answers = {}
        for row_idx in range(header_row + 1, score_row):
            q_num = ws.cell(row=row_idx, column=1).value
            correct = ws.cell(row=row_idx, column=answer_col).value
            if q_num is not None:
                try:
                    q_num = int(q_num)
                    correct_answers[q_num] = correct
                except (ValueError, TypeError):
                    pass

        return correct_answers

    def add_model_column(self, sheet_name: str, model_name: str,
                         answers: Dict[int, Any], score: float,
                         position: Optional[int] = None,
                         after_model: Optional[str] = None) -> int:
        """
        새 모델 컬럼 추가

        Args:
            sheet_name: 시트 이름
            model_name: 모델 이름
            answers: {문항번호: 답} 딕셔너리
            score: 총점
            position: 삽입할 열 번호 (1-based, None이면 마지막)
            after_model: 이 모델 다음에 삽입

        Returns:
            삽입된 열 번호
        """
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)
        score_row = self._find_score_row(sheet_name)

        model_columns = self.get_model_columns(sheet_name)

        # 이미 존재하는 모델인지 확인
        if model_name in model_columns:
            raise ValueError(f"'{model_name}' 모델이 이미 존재합니다. --update 옵션을 사용하세요.")

        # 삽입 위치 결정
        if after_model and after_model in model_columns:
            insert_col = model_columns[after_model] + 1
        elif position is not None:
            insert_col = position
        else:
            # 마지막 모델 컬럼 다음
            if model_columns:
                insert_col = max(model_columns.values()) + 1
            else:
                fixed_columns = {'문항 번호', '정답', '배점'}
                header_row = self._find_header_row(sheet_name)
                insert_col = max(
                    (
                        col_idx for col_idx in range(1, ws.max_column + 1)
                        if str(ws.cell(row=header_row, column=col_idx).value).strip()
                        in fixed_columns
                    ),
                    default=2,
                ) + 1

        # 열 삽입
        ws.insert_cols(insert_col)

        # 스타일 정의
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        red_font = Font(color='FF0000')
        red_bold_font = Font(color='FF0000', bold=True)
        purple_font = Font(color='7C3AED')
        orange_font = Font(color='C2410C')

        # 정답 데이터 가져오기
        correct_answers = self._get_correct_answers(sheet_name)

        # 헤더 설정 (볼드 + 중앙정렬)
        header_cell = ws.cell(row=header_row, column=insert_col, value=model_name)
        header_cell.font = bold_font
        header_cell.alignment = center_align

        # 답안 입력
        for row_idx in range(header_row + 1, score_row):
            q_num = ws.cell(row=row_idx, column=1).value
            if q_num is not None:
                try:
                    q_num = int(q_num)
                    cell = ws.cell(row=row_idx, column=insert_col)
                    cell.alignment = center_align

                    if q_num in answers:
                        answer = answers[q_num]
                        correct = correct_answers.get(q_num)

                        if answer == REFUSAL_ANSWER or str(answer).strip() in REFUSAL_MARKERS:
                            cell.value = "(검열)"
                            cell.font = purple_font
                        elif str(answer).strip() in PARSE_FAILED_MARKERS:
                            cell.value = "(파싱 실패)"
                            cell.font = orange_font
                        elif answer is None or answer == NO_ANSWER or answer == "":
                            cell.value = "(포기)"
                            cell.font = red_font
                        else:
                            cell.value = answer
                            # 오답 확인
                            try:
                                if int(answer) != int(correct):
                                    cell.font = red_font
                            except (ValueError, TypeError):
                                pass
                    else:
                        # answers에 없는 문항도 포기 처리
                        cell.value = "(포기)"
                        cell.font = red_font

                except (ValueError, TypeError):
                    pass

        # 총점 입력 (중앙정렬)
        score_cell = ws.cell(row=score_row, column=insert_col, value=score)
        score_cell.alignment = center_align

        # 캐시 무효화
        self._header_row_cache.pop(sheet_name, None)

        return insert_col

    def update_model_column(self, sheet_name: str, model_name: str,
                            answers: Dict[int, Any], score: float):
        """기존 모델 컬럼 업데이트"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)
        score_row = self._find_score_row(sheet_name)

        model_columns = self.get_model_columns(sheet_name)
        if model_name not in model_columns:
            raise ValueError(f"'{model_name}' 모델을 찾을 수 없습니다.")

        col_idx = model_columns[model_name]

        # 스타일 정의
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        red_font = Font(color='FF0000')
        purple_font = Font(color='7C3AED')
        orange_font = Font(color='C2410C')

        # 정답 데이터 가져오기
        correct_answers = self._get_correct_answers(sheet_name)

        # 헤더 스타일 (볼드 + 중앙정렬)
        header_cell = ws.cell(row=header_row, column=col_idx)
        header_cell.font = bold_font
        header_cell.alignment = center_align

        # 답안 업데이트
        for row_idx in range(header_row + 1, score_row):
            q_num = ws.cell(row=row_idx, column=1).value
            if q_num is not None:
                try:
                    q_num = int(q_num)
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_align
                    cell.font = Font()  # 기본 폰트로 초기화

                    if q_num in answers:
                        answer = answers[q_num]
                        correct = correct_answers.get(q_num)

                        if answer == REFUSAL_ANSWER or str(answer).strip() in REFUSAL_MARKERS:
                            cell.value = "(검열)"
                            cell.font = purple_font
                        elif str(answer).strip() in PARSE_FAILED_MARKERS:
                            cell.value = "(파싱 실패)"
                            cell.font = orange_font
                        elif answer is None or answer == NO_ANSWER or answer == "":
                            cell.value = "(포기)"
                            cell.font = red_font
                        else:
                            cell.value = answer
                            # 오답 확인
                            try:
                                if int(answer) != int(correct):
                                    cell.font = red_font
                            except (ValueError, TypeError):
                                pass
                    else:
                        cell.value = "(포기)"
                        cell.font = red_font

                except (ValueError, TypeError):
                    pass

        # 총점 업데이트 (중앙정렬)
        score_cell = ws.cell(row=score_row, column=col_idx, value=score)
        score_cell.alignment = center_align

    def save(self):
        """변경사항 저장"""
        if self.workbook:
            self.workbook.save(self.excel_path)
            print(f"저장 완료: {self.excel_path}")


class DataConverter:
    """데이터 변환 로직"""

    def __init__(self, path_mapper: PathMapper, model_mapper: ModelNameMapper):
        self.path_mapper = path_mapper
        self.model_mapper = model_mapper

    def load_questions(self, sheet_name: str) -> Dict:
        """questions.json 로드"""
        return self.path_mapper.load_questions(sheet_name)

    def excel_to_json(self, sheet_name: str, model_name: str,
                      answers: Dict[int, Any], excel_handler: ExcelHandler) -> Dict:
        """
        Excel의 한 모델 컬럼 데이터를 results_verified.json 형식으로 변환

        Args:
            sheet_name: 시트 이름
            model_name: Excel의 모델 이름
            answers: {문항번호: 답}
            excel_handler: ExcelHandler 인스턴스

        Returns:
            results_verified.json 형식의 딕셔너리
        """
        questions_data = self.load_questions(sheet_name)
        subject, section = self.path_mapper.get_subject_section(sheet_name)
        json_model_name = self.model_mapper.excel_to_json(model_name)

        results = []
        total_score = 0
        correct_count = 0
        manual_review_count = 0

        for q in questions_data['questions']:
            q_num = q['number']
            correct_answer = q['correct_answer']
            points = q['points']

            extracted = answers.get(q_num)
            extracted_normalized, answer_status = normalize_answer_value(extracted)

            is_correct = (extracted_normalized == correct_answer)
            if is_correct:
                total_score += points
                correct_count += 1
            needs_manual_review = answer_status != 'answered'
            if needs_manual_review:
                manual_review_count += 1

            results.append({
                'question_number': q_num,
                'model_name': json_model_name,
                'extracted_answer': extracted_normalized,
                'correct_answer': correct_answer,
                'is_correct': is_correct,
                'points': points,
                'needs_manual_review': needs_manual_review,
                'answer_status': answer_status,
                'provider_stop_reason': 'refusal' if answer_status == 'refusal' else None,
                'raw_response': ''
            })

        total_points = sum(q['points'] for q in questions_data['questions'])

        return {
            'subject': questions_data.get('subject', subject),
            'section': questions_data.get('section', section),
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_points': total_points,
            'total_verified': len(results),
            'correct_count': correct_count,
            'manual_review_count': manual_review_count,
            'model_scores': {json_model_name: total_score},
            'results': results
        }

    @staticmethod
    def _answer_cell(extracted: Any, answer_status: Optional[str]) -> Any:
        """
        @brief 결과 한 건의 답/상태를 Excel 셀 값으로 변환

        @param extracted 추출된 답 (센티널 포함)
        @param answer_status answered/no_answer/refusal/parse_failed
        @return Excel 셀에 기록할 값
        """
        if answer_status == 'refusal' or extracted == REFUSAL_ANSWER:
            return "(검열)"
        if answer_status == 'parse_failed':
            return "(파싱 실패)"
        if extracted == NO_ANSWER:
            return None
        return extracted

    def json_to_excel(self, json_data: Dict,
                      target_model: str = None) -> List[Tuple[str, Dict[int, Any], float]]:
        """
        results_verified.json 데이터를 Excel 형식으로 변환

        공식(v1) 점수와 별도로, lenient(v2) 결과가 v1과 다른 모델은
        "{모델명} (v2)" 컬럼을 추가로 내보내 병행 표기한다.

        Args:
            json_data: results_verified.json 데이터
            target_model: 특정 모델만 변환 (None이면 모든 모델)

        Returns:
            [(model_name, {문항번호: 답}, 총점), ...] 리스트
        """
        model_scores = json_data.get('model_scores', {})
        if not model_scores:
            raise ValueError("model_scores가 비어있습니다.")
        model_scores_lenient = json_data.get('model_scores_lenient', {})

        results_list = []

        for json_model_name, score in model_scores.items():
            # 특정 모델만 처리
            if target_model and json_model_name != target_model:
                continue

            excel_model_name = self.model_mapper.json_to_excel(json_model_name)

            # 답안 추출 (공식 v1 + 병행 v2)
            answers = {}
            lenient_answers = {}
            lenient_differs = False
            for result in json_data.get('results', []):
                if result['model_name'] != json_model_name:
                    continue
                q_num = result['question_number']
                answers[q_num] = self._answer_cell(
                    result['extracted_answer'], result.get('answer_status')
                )
                lenient_answer = result.get('lenient_answer', result['extracted_answer'])
                lenient_status = result.get('lenient_status', result.get('answer_status'))
                lenient_answers[q_num] = self._answer_cell(lenient_answer, lenient_status)
                if lenient_answers[q_num] != answers[q_num]:
                    lenient_differs = True

            results_list.append((excel_model_name, answers, score))

            lenient_score = model_scores_lenient.get(json_model_name)
            if lenient_differs and lenient_score is not None:
                results_list.append(
                    (f"{excel_model_name} (v2)", lenient_answers, lenient_score)
                )

        return results_list


class SyncManager:
    """동기화 관리"""

    def __init__(self, excel_path: Path, problems_dir: Path = None,
                 model_mapping_path: Path = None, hard_mode: bool = False,
                 benchmark_id: str = DEFAULT_BENCHMARK_ID,
                 registry_path: Path = DEFAULT_REGISTRY_PATH):
        self.excel_path = Path(excel_path)
        self.problems_dir = problems_dir or Path('problems')
        self.hard_mode = hard_mode
        self.benchmark_id = benchmark_id
        self.path_mapper = PathMapper(
            base_dir=Path('.'),
            benchmark_id=benchmark_id,
            registry_path=registry_path,
        )
        mode_config = self.path_mapper.get_mode_config(hard_mode)
        self.verified_filename = 'hard_results_verified.json' if hard_mode else 'results_verified.json'
        self.raw_results_filename = 'hard_results.json' if hard_mode else 'results.json'
        self.all_results_filename = mode_config.get(
            'results',
            'hard_all_results.json' if hard_mode else 'all_results.json',
        )
        self.token_usage_filename = 'hard_token_usage.json' if hard_mode else 'token_usage.json'
        configured_token_usage = mode_config.get('tokenUsage')
        self.token_usage_path = (
            Path(configured_token_usage)
            if configured_token_usage
            else self.problems_dir / self.token_usage_filename
        )

        self.model_mapper = ModelNameMapper(model_mapping_path)
        self.excel_handler = ExcelHandler(self.excel_path, self.path_mapper)
        self.converter = DataConverter(self.path_mapper, self.model_mapper)
        self._token_usage = self._load_token_usage()

    def _get_model_price(self, model_name: str) -> Optional[Dict[str, float]]:
        """
        @brief 특정 모델의 단가 정보를 조회한다.

        단가는 실행 시 `run_metadata.pricing`으로 기록되어 토큰 사용량 파일에
        병합되므로, import를 먼저 수행해야 조회할 수 있다.

        @param model_name 모델 이름
        @return {'input': ..., 'output': ...} 또는 없으면 None
        """
        return self._token_usage.get('models', {}).get(model_name, {}).get('price')

    def _load_token_usage(self) -> Dict:
        """토큰 사용량 파일 로드"""
        token_file = self.token_usage_path
        if not token_file.exists() and token_file.parent == Path('.'):
            legacy_token_file = self.problems_dir / token_file
            if legacy_token_file.exists():
                token_file = legacy_token_file
        self.token_usage_path = token_file
        if token_file.exists():
            with open(token_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}

    def _get_token_usage(self, model_name: str, sheet_name: str) -> Optional[Dict[str, int]]:
        """특정 모델-시트의 토큰 사용량 조회"""
        models = self._token_usage.get('models', {})
        model_data = models.get(model_name, {})
        sections = model_data.get('sections', {})

        section_data = sections.get(sheet_name)
        if section_data is None:
            return None
        return {
            'input_tokens': section_data.get('input_tokens', 0),
            'output_tokens': section_data.get('output_tokens', 0),
            'total_tokens': section_data.get(
                'total_tokens',
                section_data.get('input_tokens', 0)
                + section_data.get('output_tokens', 0),
            ),
            'cost_usd': section_data.get('cost_usd'),
        }

    def _get_verified_json_file(self, sheet_name: str) -> Optional[Path]:
        """
        @brief 시트에 대응하는 검증 결과 JSON 경로를 반환한다.

        @param sheet_name 대상 시트
        @return results_verified.json 경로 또는 매핑이 없으면 None
        """
        json_dir = self.path_mapper.sheet_to_json_path(sheet_name)
        if not json_dir:
            return None
        return json_dir / self.verified_filename

    def _load_lenient_export_data(self, sheet_name: str) -> Dict[str, Dict]:
        """
        @brief 검증 결과 JSON에서 관대 채점(v2) 정보를 읽어온다.

        공식 점수는 엄격 파서(v1)만 사용하므로 v2 값은 대시보드 병행 표기용
        부가 필드로만 내보낸다. 검증 결과 파일은 저작권상 저장소에 포함되지
        않으므로 파일이 없거나 손상된 경우 조용히 빈 값을 반환한다.

        @param sheet_name 대상 시트
        @return {모델명: {'score_lenient': float, 'correct_count_lenient': int,
                          'questions': {문항번호: {...}}}}
        """
        verified_file = self._get_verified_json_file(sheet_name)
        if not verified_file or not verified_file.exists():
            return {}

        try:
            with open(verified_file, 'r', encoding='utf-8') as file:
                json_data = json.load(file)
        except (OSError, json.JSONDecodeError) as error:
            print(f"경고: {sheet_name} 관대 채점 정보 로드 실패 - {error}")
            return {}

        if not isinstance(json_data, dict):
            return {}

        lenient_scores = json_data.get('model_scores_lenient')
        if not isinstance(lenient_scores, dict):
            lenient_scores = {}
        model_metrics = json_data.get('model_metrics')
        if not isinstance(model_metrics, dict):
            model_metrics = {}

        lenient_data: Dict[str, Dict] = {}
        for model_name, metrics in model_metrics.items():
            if not isinstance(metrics, dict):
                continue
            score = lenient_scores.get(model_name, metrics.get('lenient_score'))
            entry: Dict = {'questions': {}}
            if score is not None:
                entry['score_lenient'] = float(score)
            correct_count = metrics.get('lenient_correct_count')
            if correct_count is not None:
                entry['correct_count_lenient'] = int(correct_count)
            lenient_data[model_name] = entry

        for row in json_data.get('results', []):
            if not isinstance(row, dict):
                continue
            model_name = row.get('model_name')
            question_number = row.get('question_number')
            if model_name is None or question_number is None:
                continue
            # 관대 채점 필드가 아예 없는 행(구 스키마)은 비교 대상에서 제외한다
            if not any(
                key in row
                for key in ('lenient_answer', 'lenient_status', 'lenient_is_correct')
            ):
                continue
            entry = lenient_data.setdefault(model_name, {'questions': {}})
            entry['questions'][question_number] = {
                'lenient_answer': row.get('lenient_answer'),
                'lenient_status': row.get('lenient_status'),
                'lenient_is_correct': row.get('lenient_is_correct'),
            }

        return lenient_data

    def _merge_run_metadata(self, json_data: Dict, sheet_name: str) -> bool:
        """
        @brief 실행기의 토큰·비용 메타데이터를 대시보드용 파일에 병합한다.

        @param json_data results_verified.json 데이터
        @param sheet_name 대상 시트
        @return 토큰 사용량 파일이 변경되었는지 여부
        """
        run_metadata = json_data.get('run_metadata')
        if not isinstance(run_metadata, dict):
            return False

        models = self._token_usage.setdefault('models', {})
        changed = False
        for model_name, metadata in run_metadata.items():
            if not isinstance(metadata, dict):
                continue
            usage = metadata.get('token_usage')
            if not isinstance(usage, dict):
                continue
            input_tokens = usage.get('input_tokens')
            output_tokens = usage.get('output_tokens')
            if input_tokens is None or output_tokens is None:
                continue

            results = [
                result for result in json_data.get('results', [])
                if result.get('model_name') == model_name
            ]
            model_entry = models.setdefault(model_name, {})
            sections = model_entry.setdefault('sections', {})
            section_entry = {
                'input_tokens': int(input_tokens),
                'output_tokens': int(output_tokens),
                'total_tokens': int(input_tokens) + int(output_tokens),
                'question_count': 1 if self.hard_mode else len(results),
            }
            if metadata.get('cost_usd') is not None:
                section_entry['cost_usd'] = float(metadata['cost_usd'])
            sections[sheet_name] = section_entry

            pricing = metadata.get('pricing')
            if isinstance(pricing, dict):
                input_price = pricing.get('input_per_million')
                output_price = pricing.get('output_per_million')
                if input_price is not None and output_price is not None:
                    model_entry['price'] = {
                        'input': float(input_price),
                        'output': float(output_price),
                    }

            model_entry['provider'] = metadata.get('provider')
            model_entry['model_id'] = metadata.get('model_id')
            model_entry['total_input_tokens'] = sum(
                int(section.get('input_tokens', 0)) for section in sections.values()
            )
            model_entry['total_output_tokens'] = sum(
                int(section.get('output_tokens', 0)) for section in sections.values()
            )
            model_entry['total_tokens'] = (
                model_entry['total_input_tokens'] + model_entry['total_output_tokens']
            )
            model_entry['question_count'] = sum(
                int(section.get('question_count', 0)) for section in sections.values()
            )
            section_costs = [
                section.get('cost_usd') for section in sections.values()
                if section.get('cost_usd') is not None
            ]
            if section_costs:
                model_entry['cost_usd'] = sum(float(cost) for cost in section_costs)
            model_entry['last_updated'] = metadata.get('updated_at') or json_data.get('timestamp')
            changed = True
        return changed

    def _save_token_usage(self) -> None:
        """@brief 토큰 사용량 JSON을 원자적으로 저장한다."""
        token_path = self.token_usage_path
        token_path.parent.mkdir(parents=True, exist_ok=True)
        temporary_path = token_path.with_name(f'.{token_path.name}.tmp')
        try:
            with open(temporary_path, 'w', encoding='utf-8') as file:
                json.dump(self._token_usage, file, ensure_ascii=False, indent=2)
                file.write('\n')
            temporary_path.replace(token_path)
        finally:
            temporary_path.unlink(missing_ok=True)

    def _has_complete_model_results(self, json_data: Dict,
                                    sheet_name: str) -> bool:
        """
        @brief 결과가 모델별 전체 문항을 포함하는지 확인한다.

        중간 체크포인트를 실수로 Excel에 가져와 나머지 문항을 포기로
        확정하는 일을 막는다. 재시도를 소진해 영구 실패한 문항은 실행기가
        무응답(0점) 행으로 기록하므로 이 검사를 통과한다.

        @param json_data 가져올 결과 데이터
        @param sheet_name 대상 시트
        @return 모든 모델 결과가 완전하면 True
        """
        expected_count = self.path_mapper.get_section_config(sheet_name).get('questionCount')
        if not expected_count:
            return True

        expected_subject, expected_section = self.path_mapper.get_subject_section(sheet_name)
        if (
            json_data.get('sheet_name') != sheet_name
            or json_data.get('subject') != expected_subject
            or json_data.get('section') != expected_section
        ):
            print(
                f"시트 또는 과목 불일치: 기대={sheet_name}/{expected_subject}/{expected_section}, "
                f"JSON={json_data.get('sheet_name')}/{json_data.get('subject')}/"
                f"{json_data.get('section')}"
            )
            return False

        results = json_data.get('results', [])
        model_scores = json_data.get('model_scores', {})
        if not isinstance(model_scores, dict) or not model_scores:
            print(f"모델 점수가 비어 있습니다: {sheet_name}")
            return False

        try:
            questions = self.path_mapper.load_questions(sheet_name)['questions']
        except (FileNotFoundError, KeyError, TypeError, ValueError) as error:
            print(f"문항 메타데이터를 검증할 수 없습니다: {sheet_name} ({error})")
            return False
        expected_questions = {int(question['number']): question for question in questions}

        for model_name, declared_score in model_scores.items():
            model_results = [
                result for result in results
                if result.get('model_name') == model_name
            ]
            question_numbers = {
                result.get('question_number') for result in model_results
            }
            if len(model_results) != expected_count or question_numbers != set(
                range(1, int(expected_count) + 1)
            ):
                print(
                    f"불완전한 결과는 가져올 수 없습니다: {sheet_name} / {model_name} "
                    f"({len(model_results)}/{expected_count}문항)"
                )
                return False

            calculated_score = 0.0
            for result in model_results:
                number = result.get('question_number')
                expected = expected_questions.get(number)
                if expected is None:
                    print(f"알 수 없는 문항 결과입니다: {sheet_name} / {model_name} / {number}")
                    return False

                try:
                    metadata_matches = (
                        int(result.get('correct_answer')) == int(expected['correct_answer'])
                        and math.isclose(
                            float(result.get('points')),
                            float(expected['points']),
                            abs_tol=1e-9,
                        )
                    )
                except (TypeError, ValueError):
                    metadata_matches = False
                if not metadata_matches:
                    print(f"정답 또는 배점 불일치: {sheet_name} / {model_name} / {number}")
                    return False

                extracted = result.get('extracted_answer')
                status = result.get('answer_status')
                status_matches = (
                    (status == 'answered' and type(extracted) is int and 1 <= extracted <= 5)
                    or (status == 'no_answer' and extracted == NO_ANSWER)
                    or (status == 'refusal' and extracted == REFUSAL_ANSWER)
                    or (status == 'parse_failed' and extracted is None)
                )
                if not status_matches:
                    print(f"답안 상태 불일치: {sheet_name} / {model_name} / {number}")
                    return False

                calculated_correct = extracted == int(expected['correct_answer'])
                if result.get('is_correct') is not calculated_correct:
                    print(f"정오표 불일치: {sheet_name} / {model_name} / {number}")
                    return False
                if calculated_correct:
                    calculated_score += float(expected['points'])

            normalized_score = _normalize_numeric_value(declared_score)
            if normalized_score is None or not math.isclose(
                float(normalized_score), calculated_score, abs_tol=1e-9
            ):
                print(
                    f"선언 점수 불일치: {sheet_name} / {model_name} "
                    f"(JSON={declared_score}, 재계산={calculated_score:g})"
                )
                return False
        return True

    def _matches_import_identity(self, json_data: Dict) -> bool:
        """@brief benchmark_id와 실행 모드의 교차 가져오기를 차단한다."""
        expected_mode = 'subject' if self.hard_mode else 'question'
        benchmark_id = json_data.get('benchmark_id')
        run_mode = json_data.get('run_mode')
        if benchmark_id != self.benchmark_id:
            print(
                f"벤치마크 ID 불일치: 기대={self.benchmark_id}, JSON={benchmark_id}"
            )
            return False
        if run_mode != expected_mode:
            print(f"실행 모드 불일치: 기대={expected_mode}, JSON={run_mode}")
            return False
        return True

    def export_to_json(self, sheet_name: str, model_name: str,
                       output_path: Path = None) -> Path:
        """
        Excel -> JSON 내보내기

        Args:
            sheet_name: 시트 이름
            model_name: 모델 이름
            output_path: 출력 경로 (None이면 기본 경로)

        Returns:
            저장된 파일 경로
        """
        # 답안 추출
        answers = self.excel_handler.get_model_answers(sheet_name, model_name)

        # JSON 변환
        json_data = self.converter.excel_to_json(
            sheet_name, model_name, answers, self.excel_handler
        )
        json_data['benchmark_id'] = self.benchmark_id
        json_data['run_mode'] = 'subject' if self.hard_mode else 'question'
        json_data['sheet_name'] = sheet_name

        # 출력 경로 결정
        if output_path is None:
            json_dir = self.path_mapper.sheet_to_json_path(sheet_name)
            if not json_dir:
                raise ValueError(f"'{sheet_name}'에 대한 경로 매핑이 없습니다.")
            output_path = json_dir / self.verified_filename

        # 기존 파일이 있으면 병합
        if output_path.exists():
            with open(output_path, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)

            # 기존 모델 점수와 결과 병합
            json_model_name = self.model_mapper.excel_to_json(model_name)
            existing_data['model_scores'][json_model_name] = json_data['model_scores'][json_model_name]

            # 기존 results에서 같은 모델 결과 제거 후 추가
            existing_data['results'] = [
                r for r in existing_data['results']
                if r['model_name'] != json_model_name
            ]
            existing_data['results'].extend(json_data['results'])
            existing_data['timestamp'] = json_data['timestamp']
            existing_data['benchmark_id'] = self.benchmark_id
            existing_data['run_mode'] = 'subject' if self.hard_mode else 'question'
            existing_data['sheet_name'] = sheet_name

            json_data = existing_data

        # 저장
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)

        print(f"내보내기 완료: {output_path}")
        return output_path

    def import_from_json(self, json_path: Path,
                         position: Optional[int] = None,
                         after_model: Optional[str] = None,
                         update_existing: bool = False,
                         excel_name: Optional[str] = None,
                         base_model: Optional[str] = None) -> bool:
        """
        JSON -> Excel 가져오기

        Args:
            json_path: results_verified.json 경로
            position: 삽입할 열 위치
            after_model: 이 모델 다음에 삽입
            update_existing: 기존 데이터 업데이트 여부
            excel_name: Excel에서 사용할 모델 이름 (None이면 매핑 사용)
            base_model: Excel에서 복사해 올 베이스 모델 이름

        Returns:
            성공 여부
        """
        json_path = Path(json_path)
        if not json_path.exists():
            print(f"파일을 찾을 수 없습니다: {json_path}")
            return False

        # JSON 로드
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)

        if not self._matches_import_identity(json_data):
            return False

        # 시트 이름 결정
        sheet_name = self.path_mapper.json_to_sheet_name(json_path)
        if not sheet_name:
            print(f"경고: '{json_path}'에 대한 시트 매핑을 찾을 수 없습니다.")
            # subject/section으로 시트 이름 추론
            subject = json_data.get('subject', '')
            section = json_data.get('section', '')
            if subject and section and subject != section:
                sheet_name = f"{subject}-{section}"
            elif subject:
                sheet_name = subject
            else:
                print("시트 이름을 결정할 수 없습니다.")
                return False

        # 시트 존재 확인
        if sheet_name not in self.excel_handler.get_sheet_names():
            print(f"시트를 찾을 수 없습니다: {sheet_name}")
            return False

        if not self._has_complete_model_results(json_data, sheet_name):
            return False

        # 데이터 변환 (모든 모델 처리)
        model_data_list = self.converter.json_to_excel(json_data)

        success_count = 0
        for model_name, answers, score in model_data_list:
            # Excel 모델 이름 결정 (excel_name이 지정되면 첫 번째 모델에만 적용)
            if excel_name and success_count == 0:
                model_name = excel_name

            merged_answers = dict(answers)
            if base_model:
                existing_models = self.excel_handler.get_model_columns(sheet_name)
                if base_model not in existing_models:
                    raise ValueError(f"베이스 모델을 찾을 수 없습니다: {sheet_name} / {base_model}")
                base_answers = self.excel_handler.get_model_answers(sheet_name, base_model)
                merged_answers = dict(base_answers)
                merged_answers.update(answers)
                score = self.excel_handler.calculate_score_from_answers(sheet_name, merged_answers)

            # 기존 모델 확인
            existing_models = self.excel_handler.get_model_columns(sheet_name)

            if model_name in existing_models:
                if update_existing:
                    self.excel_handler.update_model_column(sheet_name, model_name, merged_answers, score)
                    print(f"업데이트 완료: {sheet_name} / {model_name} ({score}점)")
                    success_count += 1
                else:
                    print(f"'{model_name}' 모델이 이미 존재합니다. (건너뜀)")
            else:
                self.excel_handler.add_model_column(
                    sheet_name, model_name, merged_answers, score,
                    position=position, after_model=after_model
                )
                print(f"추가 완료: {sheet_name} / {model_name} ({score}점)")
                success_count += 1

        if success_count > 0 and self._merge_run_metadata(json_data, sheet_name):
            self._save_token_usage()
        return success_count > 0

    def import_all(self, update_existing: bool = False) -> int:
        """모든 검증 결과 JSON 가져오기"""
        count = 0
        for sheet_name in self.path_mapper.get_all_sheets():
            json_dir = self.path_mapper.sheet_to_json_path(sheet_name)
            if not json_dir:
                continue
            json_file = json_dir / self.verified_filename
            if not json_file.exists():
                continue
            if self.import_from_json(json_file, update_existing=update_existing):
                count += 1

        if count > 0:
            self.excel_handler.save()

        return count

    def export_all_sheets_to_json(self, output_path: Path = None,
                                     model_name: str = None,
                                     all_models: bool = False) -> Path:
        """
        모든 시트를 하나의 JSON 파일로 내보내기 (객체 배열 형태)

        Args:
            output_path: 출력 경로 (None이면 all_results.json)
            model_name: 특정 모델만 내보내기
            all_models: 모든 모델 내보내기

        Returns:
            저장된 파일 경로
        """
        all_data = []
        complete_models = None
        if self.path_mapper.benchmark and (
            self.path_mapper.benchmark.get('scoring', {}).get('type') == 'sum'
        ):
            required_sheets = self.path_mapper.get_all_sheets()
            workbook_sheets = set(self.excel_handler.get_sheet_names())
            if not set(required_sheets).issubset(workbook_sheets):
                complete_models = set()
            else:
                model_sets = [
                    set(self.excel_handler.get_model_columns(sheet).keys())
                    for sheet in required_sheets
                ]
                complete_models = set.intersection(*model_sets) if model_sets else set()

        for sheet_name in self.path_mapper.get_all_sheets():
            # 시트 존재 확인
            if sheet_name not in self.excel_handler.get_sheet_names():
                continue

            # 로컬 문제 또는 공개 정답 메타데이터 존재 확인
            try:
                self.path_mapper.load_questions(sheet_name)
            except (FileNotFoundError, ValueError):
                continue

            # 관대 채점(v2) 병행 표기용 부가 정보 (없으면 빈 dict)
            lenient_data = self._load_lenient_export_data(sheet_name)

            try:
                # 모델 목록 결정
                if model_name:
                    models_to_export = [model_name]
                elif all_models:
                    models_to_export = list(self.excel_handler.get_model_columns(sheet_name).keys())
                else:
                    models_to_export = list(self.excel_handler.get_model_columns(sheet_name).keys())
                # "(v2)" 병행 표기 컬럼은 Excel 전용 — 대시보드 export에서 제외
                models_to_export = [
                    model for model in models_to_export if not model.endswith(' (v2)')
                ]
                if complete_models is not None:
                    models_to_export = [
                        model for model in models_to_export if model in complete_models
                    ]

                for model in models_to_export:
                    try:
                        answers = self.excel_handler.get_model_answers(sheet_name, model)
                        json_data = self.converter.excel_to_json(
                            sheet_name, model, answers, self.excel_handler
                        )
                        # 깔끔한 형태로 재구성
                        json_model_name = list(json_data['model_scores'].keys())[0]
                        # results에서 불필요한 필드 제거
                        clean_results = [
                            {
                                'question_number': r['question_number'],
                                'extracted_answer': r['extracted_answer'],
                                'correct_answer': r['correct_answer'],
                                'is_correct': r['is_correct'],
                                'points': r['points'],
                                'answer_status': r.get('answer_status'),
                                'provider_stop_reason': r.get('provider_stop_reason'),
                            }
                            for r in json_data['results']
                        ]
                        # 관대 채점 결과가 엄격 채점과 다른 문항만 부가 표기
                        lenient_entry = lenient_data.get(json_model_name, {})
                        lenient_questions = lenient_entry.get('questions', {})
                        for clean_row in clean_results:
                            lenient_row = lenient_questions.get(
                                clean_row['question_number']
                            )
                            if not lenient_row:
                                continue
                            differs = (
                                lenient_row['lenient_is_correct'] != clean_row['is_correct']
                                or lenient_row['lenient_answer'] != clean_row['extracted_answer']
                            )
                            if differs:
                                clean_row.update(lenient_row)
                        clean_data = {
                            'benchmark_id': self.benchmark_id,
                            'run_mode': 'subject' if self.hard_mode else 'question',
                            'sheet_name': sheet_name,
                            'subject': json_data['subject'],
                            'section': json_data['section'],
                            'model_name': json_model_name,
                            'score': json_data['model_scores'][json_model_name],
                            'total_points': json_data['total_points'],
                            'correct_count': json_data['correct_count'],
                            'total_questions': json_data['total_verified'],
                        }
                        # 관대 채점(v2) 점수 병행 표기 (공식 점수는 score 유지)
                        if 'score_lenient' in lenient_entry:
                            clean_data['score_lenient'] = lenient_entry['score_lenient']
                        if 'correct_count_lenient' in lenient_entry:
                            clean_data['correct_count_lenient'] = (
                                lenient_entry['correct_count_lenient']
                            )
                        # 토큰 사용량 추가 (있는 경우에만)
                        token_usage = self._get_token_usage(json_model_name, sheet_name)
                        if token_usage:
                            clean_data['token_usage'] = token_usage
                        # 가격 정보 추가 (있는 경우에만)
                        price = self._get_model_price(json_model_name)
                        if price:
                            clean_data['price'] = price
                        clean_data['results'] = clean_results
                        all_data.append(clean_data)
                    except Exception as e:
                        print(f"경고: {sheet_name}/{model} 내보내기 실패 - {e}")

            except Exception as e:
                print(f"경고: {sheet_name} 처리 실패 - {e}")

        # 출력 경로 결정
        if output_path is None:
            output_path = Path(self.all_results_filename)

        # 저장
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(all_data, f, ensure_ascii=False, indent=2)

        print(f"내보내기 완료: {output_path} ({len(all_data)}개 항목)")
        return output_path

    def export_hard_all_sections_to_json(self, output_path: Path = None,
                                         model_name: str = None) -> Path:
        """
        hard Excel의 모든 모델 컬럼을 hard_all_results.json 형식으로 내보내기

        hard_results*.json은 개별 실행 결과 원본이므로 전체 집계의 기준으로 사용하지 않는다.
        대시보드용 전체 집계는 hard Excel을 정본으로 삼아 기존 모델을 누락하지 않는다.
        """
        return self.export_all_sheets_to_json(
            output_path=output_path,
            model_name=model_name,
            all_models=model_name is None,
        )

    def list_models(self, sheet_name: str = None) -> Dict[str, List[str]]:
        """모델 목록 반환"""
        result = {}

        if sheet_name:
            sheets = [sheet_name]
        else:
            sheets = self.excel_handler.get_sheet_names()

        for sheet in sheets:
            if sheet in self.path_mapper.SHEET_TO_JSON or sheet_name:
                try:
                    models = list(self.excel_handler.get_model_columns(sheet).keys())
                    result[sheet] = models
                except:
                    pass

        return result

    def validate(self, sheet_name: str = None) -> List[str]:
        """데이터 일관성 검증"""
        issues = []

        if sheet_name:
            sheets = [sheet_name]
        else:
            sheets = self.path_mapper.get_all_sheets()

        for sheet in sheets:
            try:
                questions = self.path_mapper.load_questions(sheet)
            except (FileNotFoundError, ValueError) as error:
                issues.append(f"[{sheet}] 문제 메타데이터 없음: {error}")
                continue

            expected_count = len(questions['questions'])
            expected_total = sum(q['points'] for q in questions['questions'])

            # Excel 데이터 확인
            try:
                model_columns = self.excel_handler.get_model_columns(sheet)
                max_score = self.excel_handler.get_max_score(sheet)
                expected_answers = {
                    int(question['number']): int(question['correct_answer'])
                    for question in questions['questions']
                }
                workbook_answers = {
                    number: int(answer)
                    for number, answer in self.excel_handler._get_correct_answers(sheet).items()
                }

                if max_score != expected_total:
                    issues.append(f"[{sheet}] 만점 불일치: Excel={max_score}, JSON={expected_total}")
                if workbook_answers != expected_answers:
                    issues.append(f"[{sheet}] Excel 정답표가 공개 정답 메타데이터와 다릅니다.")

                for model_name in model_columns:
                    answers = self.excel_handler.get_model_answers(sheet, model_name)
                    if len(answers) != expected_count:
                        issues.append(
                            f"[{sheet}] {model_name}: 문항 수 불일치 "
                            f"(Excel={len(answers)}, JSON={expected_count})"
                        )
            except Exception as e:
                issues.append(f"[{sheet}] 검증 오류: {e}")

        return issues


def main():
    parser = argparse.ArgumentParser(
        description='Excel-JSON 양방향 동기화 도구',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  # JSON -> Excel 가져오기
  python sync_data.py import --json problems/bar-exam-15/public-law/results_verified.json
  python sync_data.py import --all
  python sync_data.py import --all --hard

  # Excel -> JSON 내보내기 (단일 시트)
  python sync_data.py export --sheet 공법 --model "Claude Fable 5 (max)"
  python sync_data.py export --sheet 공법 --all-models

  # Excel -> JSON 내보내기 (모든 시트를 하나의 JSON 배열로)
  python sync_data.py export --all-sheets
  python sync_data.py export --all-sheets --output bar_exam_all_results.json
  python sync_data.py export --all-sheets --model "Claude Fable 5 (max)"  # 특정 모델만
  python sync_data.py export --all-sheets --hard

  # 모델 목록 확인
  python sync_data.py list
  python sync_data.py list --sheet 공법

  # 검증
  python sync_data.py validate
        """
    )

    parser.add_argument('--benchmark', default=DEFAULT_BENCHMARK_ID,
                        help=f'벤치마크 ID (기본값: {DEFAULT_BENCHMARK_ID})')
    parser.add_argument('--registry', default=str(DEFAULT_REGISTRY_PATH),
                        help='벤치마크 레지스트리 경로')

    subparsers = parser.add_subparsers(dest='command', help='명령')

    # Export 명령
    export_parser = subparsers.add_parser('export', help='Excel -> JSON 내보내기')
    export_parser.add_argument('--sheet', help='시트 이름 (예: 공법)')
    export_parser.add_argument('--all-sheets', action='store_true', help='모든 시트를 하나의 JSON 배열로 내보내기')
    export_parser.add_argument('--model', help='모델 이름')
    export_parser.add_argument('--all-models', action='store_true', help='모든 모델 내보내기')
    export_parser.add_argument('--output', help='출력 파일 경로')
    export_parser.add_argument('--hard', action='store_true', help='hard 전용 Excel/JSON 파일 사용')

    # Import 명령
    import_parser = subparsers.add_parser('import', help='JSON -> Excel 가져오기')
    import_parser.add_argument('--json', help='results_verified.json 경로')
    import_parser.add_argument('--all', action='store_true', help='모든 JSON 파일 가져오기')
    import_parser.add_argument('--position', type=int, help='삽입할 열 위치')
    import_parser.add_argument('--after', help='이 모델 다음에 삽입')
    import_parser.add_argument('--update', action='store_true', help='기존 데이터 업데이트')
    import_parser.add_argument('--excel-name', help='Excel에서 사용할 모델 이름')
    import_parser.add_argument('--base-model', help='기존 Excel 컬럼을 베이스로 복사한 뒤 JSON 답안만 덮어쓰기')
    import_parser.add_argument('--hard', action='store_true', help='hard 전용 Excel/JSON 파일 사용')

    # List 명령
    list_parser = subparsers.add_parser('list', help='모델 목록 확인')
    list_parser.add_argument('--sheet', help='특정 시트만')
    list_parser.add_argument('--hard', action='store_true', help='hard 전용 Excel 사용')

    # Validate 명령
    validate_parser = subparsers.add_parser('validate', help='데이터 검증')
    validate_parser.add_argument('--sheet', help='특정 시트만')
    validate_parser.add_argument('--hard', action='store_true', help='hard 전용 Excel/JSON 파일 사용')

    # 공통 옵션
    parser.add_argument('--excel', default=None,
                        help='Excel 파일 경로')
    parser.add_argument('--mapping', default='model_mapping.json',
                        help='모델 이름 매핑 파일')

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return

    hard_mode = getattr(args, 'hard', False)
    path_mapper = PathMapper(
        base_dir=Path('.'),
        benchmark_id=args.benchmark,
        registry_path=Path(args.registry),
    )
    mode_config = path_mapper.get_mode_config(hard_mode)
    configured_workbook = mode_config.get('workbook')
    default_excel_path = (
        Path(configured_workbook)
        if configured_workbook
        else (DEFAULT_HARD_EXCEL_PATH if hard_mode else DEFAULT_EXCEL_PATH)
    )
    excel_path = Path(args.excel) if args.excel else default_excel_path
    if hard_mode and not excel_path.exists():
        default_mode_config = path_mapper.get_mode_config(False)
        source_path = Path(default_mode_config.get('workbook', DEFAULT_EXCEL_PATH))
        _create_hard_excel_template(source_path, excel_path)

    # SyncManager 초기화
    sync = SyncManager(
        excel_path=excel_path,
        model_mapping_path=Path(args.mapping),
        hard_mode=hard_mode,
        benchmark_id=args.benchmark,
        registry_path=Path(args.registry),
    )

    # 명령 실행
    if args.command == 'export':
        if args.all_sheets:
            # 모든 시트를 하나의 JSON 배열로 내보내기
            output = Path(args.output) if args.output else None
            if hard_mode:
                sync.export_hard_all_sections_to_json(
                    output_path=output,
                    model_name=args.model,
                )
            else:
                sync.export_all_sheets_to_json(
                    output_path=output,
                    model_name=args.model,
                    all_models=args.all_models or (args.model is None)
                )
        elif args.sheet:
            if args.all_models:
                models = sync.list_models(args.sheet).get(args.sheet, [])
                for model in models:
                    sync.export_to_json(args.sheet, model)
            elif args.model:
                output = Path(args.output) if args.output else None
                sync.export_to_json(args.sheet, args.model, output)
            else:
                print("--model 또는 --all-models 옵션을 지정하세요.")
        else:
            print("--sheet 또는 --all-sheets 옵션을 지정하세요.")

    elif args.command == 'import':
        if args.all:
            count = sync.import_all(update_existing=args.update)
            print(f"\n총 {count}개 파일 가져오기 완료")
        elif args.json:
            success = sync.import_from_json(
                Path(args.json),
                position=args.position,
                after_model=args.after,
                update_existing=args.update,
                excel_name=args.excel_name,
                base_model=args.base_model
            )
            if success:
                sync.excel_handler.save()
        else:
            print("--json 또는 --all 옵션을 지정하세요.")

    elif args.command == 'list':
        models = sync.list_models(args.sheet)
        for sheet, model_list in models.items():
            print(f"\n[{sheet}] ({len(model_list)}개 모델)")
            for model in model_list:
                score = sync.excel_handler.get_model_score(sheet, model)
                print(f"  - {model}: {score}점")

    elif args.command == 'validate':
        issues = sync.validate(args.sheet)
        if issues:
            print("검증 결과: 문제 발견")
            for issue in issues:
                print(f"  - {issue}")
        else:
            print("검증 결과: 정상")


if __name__ == '__main__':
    main()
